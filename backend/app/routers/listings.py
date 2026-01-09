"""Listing endpoints."""
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
import csv
import io
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models import ListingResponse, ListingsResponse, HistoryEntry, ReviewUpdate
from ..database import get_connection, reset_pool

router = APIRouter()


from contextlib import contextmanager

@contextmanager
def get_connection_with_retry(max_retries: int = 3):
    """Get connection with automatic retry and pool reset on failure."""
    last_error = None
    conn = None
    
    for attempt in range(max_retries):
        try:
            # Reset pool before retry (except first attempt)
            if attempt > 0:
                reset_pool()
                time.sleep(0.5 * (attempt + 1))
            
            with get_connection() as c:
                conn = c
                yield conn
                return  # Success, exit
        except Exception as e:
            last_error = e
            print(f"[DB] Connection attempt {attempt + 1}/{max_retries} failed: {e}")
            if attempt >= max_retries - 1:
                raise  # Re-raise on final attempt

@router.get("", response_model=ListingsResponse)
async def get_listings(
    state: Optional[str] = Query(None, description="Filter by state (e.g., 'MI')"),
    states: Optional[str] = Query(None, description="Filter by multiple states (comma-separated, e.g., 'MI,CT')"),
    source: Optional[str] = Query(None, description="Filter by source (e.g., 'bizbuysell', 'bizquest')"),
    category: Optional[str] = Query(None, description="Filter by category"),
    min_price: Optional[float] = Query(None, description="Minimum asking price"),
    max_price: Optional[float] = Query(None, description="Maximum asking price"),
    min_cash_flow: Optional[float] = Query(None, description="Minimum cash flow"),
    min_revenue: Optional[float] = Query(None, description="Minimum gross revenue"),
    min_ebitda: Optional[float] = Query(None, description="Minimum EBITDA"),
    new_today: Optional[bool] = Query(None, description="Only new listings from today"),
    is_active: Optional[bool] = Query(None, description="Filter by active status (True=active, False=inactive, None=all)"),
    is_reviewed: Optional[bool] = Query(None, description="Filter by reviewed status"),
    search: Optional[str] = Query(None, description="Search in title/description"),
    sort_by: str = Query("first_seen_at", description="Sort field"),
    sort_order: str = Query("desc", description="Sort order: asc or desc"),
    page: int = Query(1, ge=1, description="Page number"),
    per_page: int = Query(50, ge=1, le=100, description="Items per page"),
):
    """Get paginated list of business listings with filters."""
    # Build query dynamically
    conditions = []
    params = []
    
    # Handle active filter (None = all, True = active only, False = inactive only)
    if is_active is True:
        conditions.append("is_active = TRUE")
    elif is_active is False:
        conditions.append("is_active = FALSE")
    
    # Handle multiple states
    if states:
        state_list = [s.strip().upper() for s in states.split(",")]
        placeholders = ", ".join(["%s"] * len(state_list))
        conditions.append(f"state IN ({placeholders})")
        params.extend(state_list)
    elif state:
        conditions.append("state = %s")
        params.append(state.upper())
    
    # Filter by source
    if source:
        conditions.append("source = %s")
        params.append(source.lower())
    
    if category:
        conditions.append("category ILIKE %s")
        params.append(f"%{category}%")
    
    if min_price is not None:
        conditions.append("asking_price >= %s")
        params.append(min_price)
    
    if max_price is not None:
        conditions.append("asking_price <= %s")
        params.append(max_price)
    
    if min_cash_flow is not None:
        conditions.append("cash_flow >= %s")
        params.append(min_cash_flow)
    
    if min_revenue is not None:
        conditions.append("gross_revenue >= %s")
        params.append(min_revenue)
    
    if min_ebitda is not None:
        conditions.append("ebitda >= %s")
        params.append(min_ebitda)
    
    if new_today:
        conditions.append("DATE(first_seen_at) = CURRENT_DATE")
    
    if is_reviewed is not None:
        conditions.append("is_reviewed = %s")
        params.append(is_reviewed)
    
    if search:
        conditions.append("(title ILIKE %s OR description ILIKE %s)")
        params.append(f"%{search}%")
        params.append(f"%{search}%")
    
    where_clause = " AND ".join(conditions) if conditions else "TRUE"
    
    # Validate sort_by
    allowed_sort = ['first_seen_at', 'last_seen_at', 'asking_price', 'cash_flow', 'gross_revenue', 'ebitda', 'title']
    if sort_by not in allowed_sort:
        sort_by = 'first_seen_at'
    
    sort_order_sql = 'DESC' if sort_order.lower() == 'desc' else 'ASC'
    
    offset = (page - 1) * per_page
    
    with get_connection_with_retry() as conn:
        with conn.cursor() as cur:
            # Get total count
            cur.execute(f"SELECT COUNT(*) FROM listings WHERE {where_clause}", params)
            total = cur.fetchone()[0]
            
            # Get listings
            query = f"""
                SELECT id, external_id, source, url, title, asking_price, cash_flow,
                       gross_revenue, ebitda, city, state, category, description, broker_name,
                       broker_company, first_seen_at, last_seen_at, last_updated_at,
                       is_active,
                       DATE(first_seen_at) = CURRENT_DATE as is_new_today,
                       last_updated_at IS NOT NULL AND last_updated_at > first_seen_at as has_price_change,
                       is_reviewed, reviewed_at, notes
                FROM listings
                WHERE {where_clause}
                ORDER BY {sort_by} {sort_order_sql}
                LIMIT %s OFFSET %s
            """
            cur.execute(query, params + [per_page, offset])
            
            rows = cur.fetchall()
    
    listings = [
        ListingResponse(
            id=row[0],
            external_id=row[1],
            source=row[2],
            url=row[3],
            title=row[4],
            asking_price=float(row[5]) if row[5] else None,
            cash_flow=float(row[6]) if row[6] else None,
            gross_revenue=float(row[7]) if row[7] else None,
            ebitda=float(row[8]) if row[8] else None,
            city=row[9],
            state=row[10],
            category=row[11],
            description=row[12],
            broker_name=row[13],
            broker_company=row[14],
            first_seen_at=row[15],
            last_seen_at=row[16],
            last_updated_at=row[17],
            is_active=row[18],
            is_new_today=row[19],
            has_price_change=row[20],
            is_reviewed=row[21] or False,
            reviewed_at=row[22],
            notes=row[23]
        )
        for row in rows
    ]
    
    pages = (total + per_page - 1) // per_page if total > 0 else 1
    
    return ListingsResponse(
        listings=listings,
        total=total,
        page=page,
        per_page=per_page,
        pages=pages
    )


@router.get("/export/csv")
async def export_csv(
    state: Optional[str] = Query(None),
    states: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    min_price: Optional[float] = Query(None),
    max_price: Optional[float] = Query(None),
    min_cash_flow: Optional[float] = Query(None),
    min_revenue: Optional[float] = Query(None),
    min_ebitda: Optional[float] = Query(None),
    is_reviewed: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
):
    """Export filtered listings to CSV."""
    conditions = ["is_active = TRUE"]
    params = []
    
    if states:
        state_list = [s.strip().upper() for s in states.split(",")]
        placeholders = ", ".join(["%s"] * len(state_list))
        conditions.append(f"state IN ({placeholders})")
        params.extend(state_list)
    
    if source:
        conditions.append("source = %s")
        params.append(source.lower())
    elif state:
        conditions.append("state = %s")
        params.append(state.upper())
    
    if min_price is not None:
        conditions.append("asking_price >= %s")
        params.append(min_price)
    
    if max_price is not None:
        conditions.append("asking_price <= %s")
        params.append(max_price)
    
    if min_cash_flow is not None:
        conditions.append("cash_flow >= %s")
        params.append(min_cash_flow)
    
    if min_revenue is not None:
        conditions.append("gross_revenue >= %s")
        params.append(min_revenue)
    
    if min_ebitda is not None:
        conditions.append("ebitda >= %s")
        params.append(min_ebitda)
    
    if is_reviewed is not None:
        conditions.append("is_reviewed = %s")
        params.append(is_reviewed)
    
    if search:
        conditions.append("(title ILIKE %s OR description ILIKE %s)")
        params.append(f"%{search}%")
        params.append(f"%{search}%")
    
    where_clause = " AND ".join(conditions)
    
    with get_connection_with_retry() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT 
                    first_seen_at, city, state, title,
                    asking_price, cash_flow, gross_revenue, ebitda,
                    description, url, is_reviewed, notes
                FROM listings
                WHERE {where_clause}
                ORDER BY first_seen_at DESC
            """, params)
            rows = cur.fetchall()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    writer.writerow([
        'Date Posted', 'City', 'State', 'Title',
        'Asking Price', 'Cash Flow/SDE', 'Revenue', 'EBITDA',
        'Description', 'URL', 'Reviewed', 'Notes'
    ])
    
    # Data rows
    for row in rows:
        writer.writerow([
            row[0].strftime('%Y-%m-%d') if row[0] else '',
            row[1] or '',
            row[2] or '',
            row[3] or '',
            f"${row[4]:,.0f}" if row[4] else '',
            f"${row[5]:,.0f}" if row[5] else '',
            f"${row[6]:,.0f}" if row[6] else '',
            f"${row[7]:,.0f}" if row[7] else '',
            (row[8] or '')[:500],
            row[9] or '',
            'Yes' if row[10] else 'No',
            row[11] or ''
        ])
    
    output.seek(0)
    
    filename = f"bizlistings_{datetime.now().strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/excel")
async def export_excel(
    state: Optional[str] = Query(None),
    states: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
    min_price: Optional[float] = Query(None),
    max_price: Optional[float] = Query(None),
    min_cash_flow: Optional[float] = Query(None),
    min_revenue: Optional[float] = Query(None),
    min_ebitda: Optional[float] = Query(None),
    is_reviewed: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
):
    """Export filtered listings to Excel (.xlsx)."""
    conditions = ["is_active = TRUE"]
    params = []
    
    if states:
        state_list = [s.strip().upper() for s in states.split(",")]
        placeholders = ", ".join(["%s"] * len(state_list))
        conditions.append(f"state IN ({placeholders})")
        params.extend(state_list)
    
    if source:
        conditions.append("source = %s")
        params.append(source.lower())
    elif state:
        conditions.append("state = %s")
        params.append(state.upper())
    
    if min_price is not None:
        conditions.append("asking_price >= %s")
        params.append(min_price)
    
    if max_price is not None:
        conditions.append("asking_price <= %s")
        params.append(max_price)
    
    if min_cash_flow is not None:
        conditions.append("cash_flow >= %s")
        params.append(min_cash_flow)
    
    if min_revenue is not None:
        conditions.append("gross_revenue >= %s")
        params.append(min_revenue)
    
    if min_ebitda is not None:
        conditions.append("ebitda >= %s")
        params.append(min_ebitda)
    
    if is_reviewed is not None:
        conditions.append("is_reviewed = %s")
        params.append(is_reviewed)
    
    if search:
        conditions.append("(title ILIKE %s OR description ILIKE %s)")
        params.append(f"%{search}%")
        params.append(f"%{search}%")
    
    where_clause = " AND ".join(conditions)
    
    with get_connection_with_retry() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT 
                    first_seen_at, city, state, source, title,
                    asking_price, cash_flow, gross_revenue, ebitda,
                    description, url, is_reviewed, notes
                FROM listings
                WHERE {where_clause}
                ORDER BY first_seen_at DESC
            """, params)
            rows = cur.fetchall()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Listings"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '$#,##0'
    
    # Headers
    headers = [
        'Date Posted', 'City', 'State', 'Source', 'Title',
        'Asking Price', 'Cash Flow/SDE', 'Revenue', 'EBITDA',
        'Description', 'URL', 'Reviewed', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_idx, row in enumerate(rows, 2):
        # Date Posted
        ws.cell(row=row_idx, column=1, value=row[0].strftime('%Y-%m-%d') if row[0] else '')
        # City
        ws.cell(row=row_idx, column=2, value=row[1] or '')
        # State
        ws.cell(row=row_idx, column=3, value=row[2] or '')
        # Source
        ws.cell(row=row_idx, column=4, value=row[3] or '')
        # Title
        ws.cell(row=row_idx, column=5, value=row[4] or '')
        # Asking Price (as number with currency format)
        price_cell = ws.cell(row=row_idx, column=6, value=row[5] if row[5] else None)
        price_cell.number_format = money_format
        # Cash Flow
        cf_cell = ws.cell(row=row_idx, column=7, value=row[6] if row[6] else None)
        cf_cell.number_format = money_format
        # Revenue
        rev_cell = ws.cell(row=row_idx, column=8, value=row[7] if row[7] else None)
        rev_cell.number_format = money_format
        # EBITDA
        ebitda_cell = ws.cell(row=row_idx, column=9, value=row[8] if row[8] else None)
        ebitda_cell.number_format = money_format
        # Description (truncated)
        ws.cell(row=row_idx, column=10, value=(row[9] or '')[:500])
        # URL (as hyperlink)
        url_cell = ws.cell(row=row_idx, column=11, value=row[10] or '')
        if row[10]:
            url_cell.hyperlink = row[10]
            url_cell.font = Font(color="0563C1", underline="single")
        # Reviewed
        ws.cell(row=row_idx, column=12, value='Yes' if row[11] else 'No')
        # Notes
        ws.cell(row=row_idx, column=13, value=row[12] or '')
    
    # Auto-adjust column widths
    column_widths = [12, 15, 8, 15, 40, 15, 15, 15, 15, 50, 40, 10, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"bizlistings_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{listing_id}", response_model=ListingResponse)
async def get_listing(listing_id: int):
    """Get a single listing by ID."""
    with get_connection_with_retry() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, external_id, source, url, title, asking_price, cash_flow,
                       gross_revenue, ebitda, city, state, category, description, broker_name,
                       broker_company, first_seen_at, last_seen_at, last_updated_at,
                       is_active,
                       DATE(first_seen_at) = CURRENT_DATE as is_new_today,
                       last_updated_at IS NOT NULL AND last_updated_at > first_seen_at as has_price_change,
                       is_reviewed, reviewed_at, notes
                FROM listings
                WHERE id = %s
            """, (listing_id,))
            
            row = cur.fetchone()
    
    if not row:
        raise HTTPException(status_code=404, detail="Listing not found")
    
    return ListingResponse(
        id=row[0],
        external_id=row[1],
        source=row[2],
        url=row[3],
        title=row[4],
        asking_price=float(row[5]) if row[5] else None,
        cash_flow=float(row[6]) if row[6] else None,
        gross_revenue=float(row[7]) if row[7] else None,
        ebitda=float(row[8]) if row[8] else None,
        city=row[9],
        state=row[10],
        category=row[11],
        description=row[12],
        broker_name=row[13],
        broker_company=row[14],
        first_seen_at=row[15],
        last_seen_at=row[16],
        last_updated_at=row[17],
        is_active=row[18],
        is_new_today=row[19],
        has_price_change=row[20],
        is_reviewed=row[21] or False,
        reviewed_at=row[22],
        notes=row[23]
    )


@router.patch("/{listing_id}/review")
async def update_review_status(listing_id: int, review: ReviewUpdate):
    """Mark a listing as reviewed or update its notes."""
    with get_connection_with_retry() as conn:
        with conn.cursor() as cur:
            # Check if listing exists
            cur.execute("SELECT id FROM listings WHERE id = %s", (listing_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Listing not found")
            
            cur.execute("""
                UPDATE listings 
                SET is_reviewed = %s, 
                    reviewed_at = CASE WHEN %s THEN NOW() ELSE reviewed_at END,
                    notes = COALESCE(%s, notes)
                WHERE id = %s
            """, (review.is_reviewed, review.is_reviewed, review.notes, listing_id))
            conn.commit()
    
    return {"status": "updated", "listing_id": listing_id}


@router.get("/{listing_id}/history", response_model=list[HistoryEntry])
async def get_listing_history(listing_id: int):
    """Get change history for a listing."""
    with get_connection_with_retry() as conn:
        with conn.cursor() as cur:
            # First check if listing exists
            cur.execute("SELECT id FROM listings WHERE id = %s", (listing_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Listing not found")
            
            cur.execute("""
                SELECT id, changed_at, change_type, old_values, new_values, changed_fields
                FROM listing_history
                WHERE listing_id = %s
                ORDER BY changed_at DESC
            """, (listing_id,))
            
            rows = cur.fetchall()
    
    return [
        HistoryEntry(
            id=row[0],
            changed_at=row[1],
            change_type=row[2],
            old_values=row[3],
            new_values=row[4],
            changed_fields=row[5]
        )
        for row in rows
    ]
